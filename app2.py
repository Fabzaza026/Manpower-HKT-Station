import io
import re
import datetime
from pathlib import Path
import openpyxl
import pandas as pd
import numpy as np
import streamlit as st

# ==========================================
# 1. PAGE CONFIGURATION & HIGH-CONTRAST STYLING
# ==========================================
st.set_page_config(
    page_title="Daily Manpower Allocation System",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for High-Contrast Light Theme & Dark Text
st.markdown("""
    <style>
    /* Global Background & Dark Text for Readability */
    .main {
        background-color: #F8FAFC;
    }
    
    /* Custom Headers with Dark Text */
    .main-title { 
        font-size: 28px; 
        font-weight: 800; 
        color: #0F172A; 
        margin-bottom: 5px; 
        letter-spacing: -0.5px;
    }
    .sub-title { 
        font-size: 14px; 
        color: #334155; 
        margin-bottom: 25px; 
        font-weight: 500;
    }
    
    /* Force Dark Text inside Streamlit elements */
    .stMarkdown, p, span, label {
        color: #1E293B !important;
    }
    
    /* Sidebar Styling */
    [data-testid="stSidebar"] {
        background-color: #F1F5F9;
        border-right: 1px solid #E2E8F0;
    }
    [data-testid="stSidebar"] .stMarkdown, [data-testid="stSidebar"] label {
        color: #0F172A !important;
        font-weight: 600;
    }
    </style>
""", unsafe_allow_html=True)

# App Header
st.markdown('<div class="main-title">✈️ Daily Manpower Allocation & Rostering System</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Automated shift balancing, license verification, strict LAE workload capping, CERT weighting, TR specific rule (1 Mech, 1 LAE), automatic TG 2-Mech requirement, and rotating "Work Force" assignment.</div>', unsafe_allow_html=True)

# ==========================================
# 2. HELPER FUNCTIONS & PARSERS
# ==========================================
LAE_NAME_LIST = [
    "SOMS", "RIT", "KAM", "JAI", "THA", "SUPP", "AUK", 
    "REP", "YON", "SUPA", "PEE", "NAT", "SOMK", "PUD", "NAP", "PRAK"
]
LAE_FLEX_MECH_LIST = ["PRAK", "AUK", "NAP", "SUPP", "PEE", "PUD"]

def calculate_flight_weight(contract_type, airline=""):
    airline_clean = str(airline).strip().upper()
    if airline_clean == 'TG':
        return 1.0
    if pd.isna(contract_type):
        return 1.0
    c_type = str(contract_type).strip().upper()
    if c_type in ['FULL', 'ASSIST', 'CERT']:
        return 1.0
    elif c_type in ['ON CALL', 'ONCALL']:
        return 0.5
    return 1.0

def calculate_lae_weight(contract_type, airline=""):
    airline_clean = str(airline).strip().upper()
    c_type = str(contract_type).strip().upper() if pd.notna(contract_type) else ""
    if airline_clean in ['TG', 'TR'] or c_type in ['ON CALL', 'ONCALL']:
        return 0.5
    return calculate_flight_weight(contract_type, airline)

def get_staff_effective_weight(code, name, contract_type, airline, role_col_type):
    code_clean = str(code).strip().upper()
    name_clean = str(name).strip().upper()
    is_flex = code_clean in LAE_FLEX_MECH_LIST or any(f_code in name_clean for f_code in LAE_FLEX_MECH_LIST)
    if is_flex:
        return calculate_flight_weight(contract_type, airline)
    if role_col_type == 'LAE':
        return calculate_lae_weight(contract_type, airline)
    else:
        return calculate_flight_weight(contract_type, airline)

def parse_time_str(val):
    if pd.isna(val) or val == 'None' or not str(val).strip():
        return None
    val_str = str(val).strip().split('.')[0]
    if len(val_str) == 3:
        val_str = '0' + val_str
    if len(val_str) == 4 and val_str.isdigit():
        return datetime.time(int(val_str[:2]), int(val_str[2:]))
    try:
        return pd.to_datetime(val_str).time()
    except Exception:
        return None

def get_full_datetime_range(f_sta, f_std):
    if not (f_sta and f_std):
        return None, None
    dummy_date = datetime.date(2026, 1, 1)
    dt_sta = datetime.datetime.combine(dummy_date, f_sta)
    dt_std = datetime.datetime.combine(dummy_date, f_std)
    if dt_std <= dt_sta:
        dt_std += datetime.timedelta(days=1)
    return dt_sta, dt_std

def get_shift_datetime_range(t_in, t_out):
    if not (t_in and t_out):
        return None, None
    dummy_date = datetime.date(2026, 1, 1)
    dt_in = datetime.datetime.combine(dummy_date, t_in)
    dt_out = datetime.datetime.combine(dummy_date, t_out)
    if dt_out <= dt_in:
        dt_out += datetime.timedelta(days=1)
    return dt_in, dt_out

def has_sufficient_gap(f1_start, f1_end, f2_start, f2_end, min_gap_minutes=10):
    dt1_sta, dt1_std = get_full_datetime_range(f1_start, f1_end)
    dt2_sta, dt2_std = get_full_datetime_range(f2_start, f2_end)
    if not (dt1_sta and dt1_std and dt2_sta and dt2_std):
        return True
    if max(dt1_sta, dt2_sta) < min(dt1_std, dt2_std):
        return False
    gap_delta = datetime.timedelta(minutes=min_gap_minutes)
    if dt1_std <= dt2_sta and (dt1_std + gap_delta > dt2_sta):
        return False
    if dt2_std <= dt1_sta and (dt2_std + gap_delta > dt1_sta):
        return False
    return True

def is_within_shift(f_sta, f_std, t_in, t_out):
    if not (f_sta and f_std and t_in and t_out):
        return True  
    f_start_dt, f_end_dt = get_full_datetime_range(f_sta, f_std)
    shift_in_dt, shift_out_dt = get_shift_datetime_range(t_in, t_out)
    if not (f_start_dt and f_end_dt and shift_in_dt and shift_out_dt):
        return True
    if shift_in_dt <= f_start_dt and f_end_dt <= shift_out_dt:
        return True
    shift_in_dt_next = shift_in_dt + datetime.timedelta(days=1)
    shift_out_dt_next = shift_out_dt + datetime.timedelta(days=1)
    if shift_in_dt_next <= f_start_dt and f_end_dt <= shift_out_dt_next:
        return True
    shift_in_dt_prev = shift_in_dt - datetime.timedelta(days=1)
    shift_out_dt_prev = shift_out_dt - datetime.timedelta(days=1)
    if shift_in_dt_prev <= f_start_dt and f_end_dt <= shift_out_dt_prev:
        return True
    return False

def check_privilege(staff_code, staff_privileges, ac_type):
    code_clean = str(staff_code).strip().upper()
    ac_type_clean = str(ac_type).strip().upper()
    if code_clean in ["SUPP", "THA"] and ac_type_clean in ["A321", "A32N"]:
        return False
    if code_clean in ["JAI", "SOMS"] and ac_type_clean == "B7M8":
        return False
    if pd.isna(staff_privileges) or not str(staff_privileges).strip():
        return False
    privs = [p.strip().upper() for p in str(staff_privileges).split(',')]
    if 'ALL' in privs:
        return True
    model_mapping = {
        'A320': ['A320', 'A32N', 'A321', 'A319', 'A320NEO'],
        'B737': ['B737', 'B738', 'B7M8', 'B739', 'B737MAX'],
        'B777': ['B777', 'B77W', 'B772', 'B773'],
        'A350': ['A350', 'A359', 'A3510'],
        'B787': ['B787', 'B788', 'B789', 'B7810'],
        'A330': ['A330', 'A332', 'A333', 'A339'],
        'B767': ['B767', 'B763']
    }
    for license_family, models in model_mapping.items():
        if any(m in ac_type_clean for m in models) or ac_type_clean in models:
            if license_family in privs:
                return True
    return any(p in ac_type_clean or ac_type_clean in p for p in privs)

def check_customer_authorize(staff_customers, flight_airline):
    if pd.isna(staff_customers) or not str(staff_customers).strip() or str(staff_customers).strip().upper() in ['NONE', 'ALL', '']:
        return True
    airline_clean = str(flight_airline).strip().upper()
    cust_list = [c.strip().upper() for c in str(staff_customers).split(',')]
    return 'ALL' in cust_list or airline_clean in cust_list or any(c in airline_clean for c in cust_list)

def get_latest_schedule_sheet(sheet_names):
    if not sheet_names:
        return None
    def extract_sheet_date(sheet_name):
        text = str(sheet_name)
        match = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{8})', text)
        if not match:
            return pd.NaT
        try:
            return pd.to_datetime(match.group(1), errors='coerce')
        except Exception:
            return pd.NaT

    dated_sheets = [(sheet, extract_sheet_date(sheet)) for sheet in sheet_names]
    valid_dates = [(sheet, ts) for sheet, ts in dated_sheets if pd.notna(ts)]
    if valid_dates:
        return max(valid_dates, key=lambda item: item[1])[0]
    return sheet_names[-1]


def parse_schedule_excel(uploaded_file, sheet_name, license_df=None):
    uploaded_file.seek(0)
    df_raw = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
    
    flights = []
    for idx in range(2, len(df_raw)):
        if 54 <= idx <= 61:
            continue
        row = df_raw.iloc[idx]
        row_str_combined = " ".join([str(val).upper() for val in row.values if pd.notna(val)])
        if "CNL" in row_str_combined:
            continue
            
        log_no = str(row[0]).strip() if pd.notna(row[0]) else ""
        airline = str(row[1]).strip() if pd.notna(row[1]) else ""
        flt_no = str(row[2]).strip() if pd.notna(row[2]) else ""
        sta = str(row[4]).strip() if pd.notna(row[4]) else ""
        responsible = str(row[5]).strip() if pd.notna(row[5]) and str(row[5]) != 'None' else ""
        std = str(row[6]).strip() if pd.notna(row[6]) else ""
        ac_type = str(row[8]).strip() if pd.notna(row[8]) else ""
        contract = str(row[14]).strip() if pd.notna(row[14]) else "FULL"
        existing_lae = str(row[11]).strip() if pd.notna(row[11]) and str(row[11]) != 'None' else ""
        existing_mech1 = str(row[9]).strip() if pd.notna(row[9]) and str(row[9]) != 'None' else ""
        existing_mech2 = str(row[10]).strip() if pd.notna(row[10]) and str(row[10]) != 'None' else ""
        
        if not flt_no or flt_no.upper() == 'NONE' or airline.upper() == 'A/L':
            continue
            
        flights.append({
            "Row_Idx": idx,
            "LOG_No": log_no,
            "Flight_ID": f"{airline}{flt_no}",
            "Airline": airline,
            "Flight_No": flt_no,
            "Responsible": responsible,
            "A/C Type": ac_type,
            "Contract": contract,
            "STA": sta,
            "STD": std,
            "LAE": existing_lae,
            "Mech_1": existing_mech1,
            "Mech_2": existing_mech2,
        })
    
    staff_members = []
    for idx in range(1, len(df_raw)):
        if 54 <= idx <= 61:
            continue
        row = df_raw.iloc[idx]
        col19_name = str(row[19]).strip() if pd.notna(row[19]) else ""
        pers_id = str(row[20]).strip() if pd.notna(row[20]) else ""
        col21_skd = str(row[21]).strip().upper() if pd.notna(row[21]) else ""
        cd_val = row[22] if pd.notna(row[22]) else None
        existing_assignment = str(row[25]).strip() if pd.notna(row[25]) and str(row[25]) != 'None' else ""
        code = str(row[27]).strip() if pd.notna(row[27]) and str(row[27]) != 'None' else ""
        
        if col19_name and col19_name.upper() != 'NAME':
            if not code:
                clean_name = re.sub(r'[^a-zA-Z]', '', col19_name)
                code = clean_name[:4].upper() if clean_name else "UNK"
            
            privileges = "None"
            customers = "ALL"
            
            if license_df is not None and not license_df.empty:
                match = license_df[license_df['Personal ID'].astype(str).str.strip() == pers_id]
                if not match.empty:
                    if 'Privileges' in match.columns:
                        privileges = str(match.iloc[0]['Privileges'])
                    elif 'Privilages' in match.columns:
                        privileges = str(match.iloc[0]['Privilages'])
                    for cust_col in ['Customer', 'Customer_Authorize', 'Customers', 'Customer Authorize']:
                        if cust_col in match.columns and pd.notna(match.iloc[0][cust_col]):
                            customers = str(match.iloc[0][cust_col])
                            break
            name_upper = col19_name.upper()
            code_upper = code.upper()
            name_tokens = re.findall(r'[A-Z0-9]+', name_upper)
            
            is_lae = False
            for lae_code in LAE_NAME_LIST:
                if code_upper == lae_code or lae_code in name_tokens:
                    is_lae = True
                    break
                    
            current_role = "LAE" if is_lae else "Mech"
            time_in = parse_time_str(row[23])
            time_out = parse_time_str(row[24])
            shift = col21_skd
            staff_members.append({
                "Row_Idx": idx,
                "Personal ID": pers_id,
                "Code": code,
                "Name": col19_name,
                "Role": current_role,
                "Shift": shift,
                "CD": cd_val,
                "Time_IN": time_in,
                "Time_OUT": time_out,
                "Assignment": existing_assignment,
                "Privileges": privileges,
                "Customer": customers
            })
            
    day_seed = 0
    date_match = re.search(r'(\d+)', str(sheet_name))
    if date_match:
        day_seed = int(date_match.group(1))
    else:
        day_seed = abs(hash(str(sheet_name))) % 1000
    d_candidates = [m for m in staff_members if str(m['Shift']).strip().upper() == 'D' and 
                    (str(m['Role']).strip().upper() == 'MECH' or str(m['Code']).strip().upper() in LAE_FLEX_MECH_LIST or any(f in str(m['Name']).upper() for f in LAE_FLEX_MECH_LIST))]
    
    n_candidates = [m for m in staff_members if str(m['Shift']).strip().upper() == 'N' and 
                    (str(m['Role']).strip().upper() == 'MECH' or str(m['Code']).strip().upper() in LAE_FLEX_MECH_LIST or any(f in str(m['Name']).upper() for f in LAE_FLEX_MECH_LIST))]
    selected_d_code = None
    selected_n_code = None
    if d_candidates:
        d_candidates.sort(key=lambda x: x['Code'])
        selected_idx = day_seed % len(d_candidates)
        selected_d_code = d_candidates[selected_idx]['Code']
    if n_candidates:
        n_candidates.sort(key=lambda x: x['Code'])
        selected_idx = (day_seed + 1) % len(n_candidates)
        selected_n_code = n_candidates[selected_idx]['Code']
    for s_member in staff_members:
        s_code = str(s_member['Code']).strip().upper()
        s_shift = str(s_member['Shift']).strip().upper()
        
        if s_shift == 'D' and s_code == selected_d_code:
            s_member['Assignment'] = "Work Force"
        elif s_shift == 'N' and s_code == selected_n_code:
            s_member['Assignment'] = "Work Force"
    flights_df = pd.DataFrame(flights)
    staff_df = pd.DataFrame(staff_members).drop_duplicates(subset=['Name']) if staff_members else pd.DataFrame()
    return flights_df, staff_df

def recalculate_manual_workload(flights_df, staff_df):
    workload = {row['Code']: 0.0 for _, row in staff_df.iterrows()}
    code_to_flights = {row['Code']: [] for _, row in staff_df.iterrows()}
    code_to_name = {row['Code']: row['Name'] for _, row in staff_df.iterrows()}
    for _, row in staff_df.iterrows():
        code = row['Code']
        current_assign = str(row.get('Assignment', '')).strip()
        if current_assign == "Work Force":
            code_to_flights[code] = ["Work Force"]
    for _, row in flights_df.iterrows():
        flt_id = row['Flight_ID']
        airline = row['Airline']
        contract = row['Contract']
        
        for col in ['LAE', 'Mech_1', 'Mech_2']:
            code = str(row.get(col, '')).strip().upper()
            if code and code != 'NONE':
                if code in workload:
                    s_name = code_to_name.get(code, "")
                    w = get_staff_effective_weight(code, s_name, contract, airline, col)
                    workload[code] += w
                    
                    if "Work Force" in code_to_flights[code]:
                        code_to_flights[code] = [f for f in code_to_flights[code] if f != "Work Force"]
                    if flt_id not in code_to_flights[code]:
                        code_to_flights[code].append(flt_id)
                        
    staff_updated = staff_df.copy()
    for idx, row in staff_updated.iterrows():
        code = row['Code']
        if code in code_to_flights and code_to_flights[code]:
            staff_updated.at[idx, 'Assignment'] = ", ".join(code_to_flights[code])
        else:
            staff_updated.at[idx, 'Assignment'] = ""
            
    return staff_updated, workload

def auto_allocate_manpower(flights_df, staff_df, balance_shifts=True):
    allocated_flights = flights_df.copy()
    allocated_staff = staff_df.copy()
    
    off_shifts = ['O', 'O1', 'O2']
    valid_lae_contracts = ['FULL', 'ON CALL', 'ONCALL', 'CERT']
    excluded_lae_flights = ['FM857', 'FM858', 'FM831', 'FM832']
    
    workload = {row['Code']: 0.0 for _, row in allocated_staff.iterrows()}
    assignments = {row['Code']: [] for _, row in allocated_staff.iterrows()}
    
    workforce_codes = set()
    for _, s_row in allocated_staff.iterrows():
        if str(s_row.get('Assignment', '')).strip() == "Work Force":
            workforce_codes.add(str(s_row['Code']).strip().upper())

    def shift_workload(shift):
        shift_clean = str(shift).strip().upper()
        return sum(
            workload[code]
            for code, staff_row in allocated_staff.set_index('Code').iterrows()
            if str(staff_row['Shift']).strip().upper() == shift_clean
        )
            
    def count_assigned_flights(s_code):
        flt_list = assignments.get(s_code, [])
        return len(flt_list)
        
    for idx, row in allocated_flights.iterrows():
        airline = str(row['Airline']).strip().upper()
        flt_no = str(row['Flight_No']).strip().upper()
        flt_id = f"{airline}{flt_no}".replace(" ", "")
        ac_type = str(row['A/C Type']).upper()
        contract = str(row['Contract']).strip().upper()
        responsible = str(row.get('Responsible', '')).strip().upper()
        
        f_in = parse_time_str(row['STA'])
        f_std = parse_time_str(row['STD'])
        
        # --- Allocate LAE ---
        is_excluded_flight = any(ex in flt_id for ex in excluded_lae_flights)
        
        if ((contract in valid_lae_contracts) or (airline == 'TR')) and (not is_excluded_flight):
            if not str(row['LAE']).strip():
                lae_candidates = allocated_staff[
                    (allocated_staff['Role'] == 'LAE') & 
                    (~allocated_staff['Shift'].isin(off_shifts)) &
                    (~allocated_staff['Code'].str.upper().isin(workforce_codes))
                ]
                
                if responsible and responsible != 'NONE':
                    lae_resp_filter = lae_candidates['Code'].str.contains(responsible, case=False, na=False) | \
                                      lae_candidates['Name'].str.contains(responsible, case=False, na=False)
                    if lae_resp_filter.any():
                        lae_candidates = lae_candidates[lae_resp_filter]
                
                eligible_lae = []
                for _, lae in lae_candidates.iterrows():
                    code = lae['Code']
                    s_name = lae['Name']
                    code_clean = str(code).strip().upper()
                    
                    # --- เพิ่มเงื่อนไข ห้ามทำงานร่วมกันระหว่าง JAI และกลุ่ม [SUPP, PRAK, PUD, ITT, AUK] ---
                    restricted_group = ['SUPP', 'PRAK', 'PUD', 'ITT', 'AUK']
                    current_flight_crew = [
                        str(allocated_flights.at[idx, 'LAE']).strip().upper(),
                        str(allocated_flights.at[idx, 'Mech_1']).strip().upper(),
                        str(allocated_flights.at[idx, 'Mech_2']).strip().upper()
                    ]
                    if code_clean == 'JAI' and any(r in current_flight_crew for r in restricted_group):
                        continue
                    if code_clean in restricted_group and 'JAI' in current_flight_crew:
                        continue
                    # ----------------------------------------------------------------------------------

                    if airline == 'TG' and code_clean in ['PEE', 'SUPP']:
                        continue
                    privs = str(lae['Privileges'])
                    cust_auth = str(lae.get('Customer', 'ALL'))
                    t_in = lae['Time_IN']
                    t_out = lae['Time_OUT']
                    
                    lae_weight = get_staff_effective_weight(code, s_name, contract, airline, 'LAE')
                    
                    cd_val = lae['CD']
                    if cd_val is not None and not pd.isna(cd_val):
                        try:
                            if int(float(cd_val)) in [13, 21, 64, 41]:
                                continue
                        except (ValueError, TypeError):
                            pass
                    
                    has_priv = check_privilege(code, privs, ac_type)
                    has_cust = check_customer_authorize(cust_auth, airline)
                    
                    has_cap = (workload[code] + lae_weight) <= 4.0
                    has_gap = all(
                        has_sufficient_gap(f_in, f_std, s_in, s_out, min_gap_minutes=10) 
                        for s_in, s_out in assignments[code]
                    )
                    in_shift = is_within_shift(f_in, f_std, t_in, t_out)
                    
                    if has_priv and has_cust and has_cap and has_gap and in_shift:
                        eligible_lae.append(code)
                        
                if eligible_lae:
                    if balance_shifts:
                        eligible_lae.sort(key=lambda code: (
                            shift_workload(allocated_staff.loc[allocated_staff['Code'] == code, 'Shift'].iloc[0]),
                            workload[code]
                        ))
                    selected_code = eligible_lae[0]
                    allocated_flights.at[idx, 'LAE'] = selected_code
                    
                    sel_row = allocated_staff[allocated_staff['Code'] == selected_code].iloc[0]
                    sel_weight = get_staff_effective_weight(selected_code, sel_row['Name'], contract, airline, 'LAE')
                    workload[selected_code] += sel_weight
                    
                    if f_in and f_std: 
                        assignments[selected_code].append((f_in, f_std))
                else:
                    allocated_flights.at[idx, 'LAE'] = ""
        else:
            allocated_flights.at[idx, 'LAE'] = ""
            
        # --- Allocate Mechanics ---
        if airline == 'TG':
            req_mech_count = 2
        elif airline == 'TR':
            req_mech_count = 1
        else:
            req_mech_count = 2 if contract in ['FULL', 'ASSIST', 'CERT'] else 1
            
        mech_candidates = allocated_staff[
            (~allocated_staff['Shift'].isin(off_shifts)) &
            (~allocated_staff['Code'].str.upper().isin(workforce_codes))
        ]
        
        if responsible and responsible != 'NONE':
            mech_resp_filter = mech_candidates['Code'].str.contains(responsible, case=False, na=False) | \
                               mech_candidates['Name'].str.contains(responsible, case=False, na=False)
            if mech_resp_filter.any():
                mech_candidates = mech_candidates[mech_resp_filter]
        
        for mech_col in ['Mech_1', 'Mech_2'][:req_mech_count]:
            if not str(row[mech_col]).strip():
                eligible_mech = []
                
                for _, mech in mech_candidates.iterrows():
                    code = mech['Code']
                    role = mech['Role']
                    s_name = mech['Name']
                    name_upper = str(s_name).upper()
                    t_in = mech['Time_IN']
                    t_out = mech['Time_OUT']
                    m_shift = str(mech['Shift']).strip().upper()
                    
                    code_clean = str(code).strip().upper()
                    
                    # --- เพิ่มเงื่อนไข ห้ามทำงานร่วมกันระหว่าง JAI และกลุ่ม [SUPP, PRAK, PUD, ITT, AUK] ---
                    restricted_group = ['SUPP', 'PRAK', 'PUD', 'ITT', 'AUK']
                    current_flight_crew = [
                        str(allocated_flights.at[idx, 'LAE']).strip().upper(),
                        str(allocated_flights.at[idx, 'Mech_1']).strip().upper(),
                        str(allocated_flights.at[idx, 'Mech_2']).strip().upper()
                    ]
                    if code_clean == 'JAI' and any(r in current_flight_crew for r in restricted_group):
                        continue
                    if code_clean in restricted_group and 'JAI' in current_flight_crew:
                        continue
                    # ----------------------------------------------------------------------------------

                    is_flex_mech = code_clean in LAE_FLEX_MECH_LIST or any(f_code in name_upper for f_code in LAE_FLEX_MECH_LIST)
                    
                    if role == 'LAE' and not is_flex_mech:
                        continue
                        
                    if code == allocated_flights.at[idx, 'LAE']:
                        continue
                    if mech_col == 'Mech_2' and code == allocated_flights.at[idx, 'Mech_1']:
                        continue
                    mech_weight = get_staff_effective_weight(code, s_name, contract, airline, mech_col)
                    cd_val = mech['CD']
                    if cd_val is not None and not pd.isna(cd_val):
                        try:
                            if int(float(cd_val)) in [13, 21, 64, 41]:
                                continue
                        except (ValueError, TypeError):
                            pass
                            
                    flight_limit_ok = count_assigned_flights(code) < 3
                    has_cap = (workload[code] + mech_weight) <= 4.0 and flight_limit_ok
                    has_gap = all(
                        has_sufficient_gap(f_in, f_std, s_in, s_out, min_gap_minutes=10) 
                        for s_in, s_out in assignments[code]
                    )
                    in_shift = is_within_shift(f_in, f_std, t_in, t_out)
                    
                    if has_cap and has_gap and in_shift:
                        eligible_mech.append((code, m_shift, workload[code]))
                
                if eligible_mech:
                    if balance_shifts:
                        eligible_mech.sort(key=lambda x: (shift_workload(x[1]), x[2]))
                    else:
                        eligible_mech.sort(key=lambda x: x[2])
                    selected_code = eligible_mech[0][0]
                    
                    allocated_flights.at[idx, mech_col] = selected_code
                    
                    sel_row = allocated_staff[allocated_staff['Code'] == selected_code].iloc[0]
                    sel_weight = get_staff_effective_weight(selected_code, sel_row['Name'], contract, airline, mech_col)
                    workload[selected_code] += sel_weight
                    
                    if f_in and f_std: 
                        assignments[selected_code].append((f_in, f_std))
                else:
                    allocated_flights.at[idx, mech_col] = ""
                    
    allocated_staff, workload = recalculate_manual_workload(allocated_flights, allocated_staff)
    return allocated_flights, allocated_staff, workload

def export_to_excel(original_file, sheet_name, flights_df, staff_df):
    original_file.seek(0)
    wb = openpyxl.load_workbook(original_file)
    wb.security = None
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    ws.protection.disable()
    for _, row in flights_df.iterrows():
        r_idx = row['Row_Idx'] + 1
        ws.cell(row=r_idx, column=10, value=row.get('Mech_1', ''))
        ws.cell(row=r_idx, column=11, value=row.get('Mech_2', ''))
        ws.cell(row=r_idx, column=12, value=row.get('LAE', ''))
    for _, row in staff_df.iterrows():
        r_idx = row['Row_Idx'] + 1
        ws.cell(row=r_idx, column=26, value=row.get('Assignment', ''))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 3. SIDEBAR & FILE UPLOADERS
# ==========================================
st.sidebar.markdown("### 📂 Data Management")
st.sidebar.markdown("---")
st.sidebar.markdown("📌 **Step 1:** Loading master license data.")
master_license_path = Path(__file__).with_name("manpower_license_data 2.xlsx")
license_df = None
try:
    excel_lic = pd.ExcelFile(master_license_path)
    target_sheet = "Manpower_Licenses" if "Manpower_Licenses" in excel_lic.sheet_names else excel_lic.sheet_names[0]
    license_df = pd.read_excel(master_license_path, sheet_name=target_sheet)
    st.sidebar.success("✅ Master License loaded from manpower_license_data 2.xlsx!")
except Exception as e:
    st.sidebar.error(f"❌ Master License error: {e}")

st.sidebar.markdown("---")
st.sidebar.markdown("📌 **Step 2:** Upload Daily Flight Schedule.")
schedule_file = st.sidebar.file_uploader("Daily Schedule (.xlsx)", type=["xlsx", "xls"])
flights_df, staff_df = pd.DataFrame(), pd.DataFrame()
selected_sheet = None
if schedule_file is not None:
    try:
        excel_file = pd.ExcelFile(schedule_file)
        sheet_names = excel_file.sheet_names
        selected_sheet = get_latest_schedule_sheet(sheet_names)
        if selected_sheet is not None:
            flights_df, staff_df = parse_schedule_excel(schedule_file, selected_sheet, license_df)
            st.sidebar.success(f"✅ Latest sheet '{selected_sheet}' loaded automatically!")
        else:
            st.sidebar.warning("⚠️ No sheets found in the uploaded workbook.")
    except Exception as e:
        st.sidebar.error(f"❌ Error: {e}")
else:
    st.sidebar.info("💡 Tip: Upload both files to initialize the dashboard.")

if 'flights_data' not in st.session_state or schedule_file:
    st.session_state.flights_data = flights_df
if 'staff_data' not in st.session_state or schedule_file:
    st.session_state.staff_data = staff_df
if 'workload_summary' not in st.session_state:
    st.session_state.workload_summary = {}

# ==========================================
# 4. MAIN DASHBOARD TABS
# ==========================================
tab1, tab2, tab3 = st.tabs([
    "✈️ Flight Allocation", 
    "👥 Staff Roster & Authorizations", 
    "📊 Workload & Shift Balance Analysis"
])

# --- TAB 1: FLIGHT ALLOCATION ---
with tab1:
    st.markdown("### 🛫 Flight Manpower Assignment Table")
    st.markdown("Execute automated matching or manually inspect and override entries below.")
    
    col_a, col_b = st.columns([1, 2])
    with col_a:
        balance_shift_option = st.checkbox("⚖️ Enable Shift Workload Balance (D vs N)", value=True, help="Balances workload distribution between Day (D) and Night (N) shifts during auto-allocation.")
        if st.button("🚀 Run Auto-Allocation", type="primary", use_container_width=True):
            if not st.session_state.flights_data.empty and not st.session_state.staff_data.empty:
                res_flights, res_staff, wl_summary = auto_allocate_manpower(
                    st.session_state.flights_data, 
                    st.session_state.staff_data,
                    balance_shifts=balance_shift_option
                )
                st.session_state.flights_data = res_flights
                st.session_state.staff_data = res_staff
                st.session_state.workload_summary = wl_summary
                st.success("✨ Automated allocation completed successfully!")
            else:
                st.warning("⚠️ Please upload a Daily Schedule before executing.")
    with col_b:
        if schedule_file is not None and not st.session_state.flights_data.empty:
            excel_bytes = export_to_excel(
                schedule_file, 
                selected_sheet, 
                st.session_state.flights_data, 
                st.session_state.staff_data
            )
            st.download_button(
                label="📥 Download Updated Excel Report",
                data=excel_bytes,
                file_name=f"Allocated_{schedule_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                use_container_width=True
            )
            
    st.markdown("---")
    if not st.session_state.flights_data.empty:
        display_df = st.session_state.flights_data.copy()
        display_df['Flight_Points'] = display_df.apply(
            lambda r: calculate_flight_weight(r['Contract'], r['Airline']), axis=1
        )
        cols_to_show = ["LOG_No", "Flight_ID", "Airline", "Flight_No", "Responsible", "A/C Type", "Contract", "STA", "STD", "Flight_Points", "LAE", "Mech_1", "Mech_2"]
        cols_exist = [c for c in cols_to_show if c in display_df.columns]
        
        edited_flights = st.data_editor(
            display_df[cols_exist],
            num_rows="dynamic",
            use_container_width=True,
            key="flight_editor",
            column_config={
                "Contract": st.column_config.SelectboxColumn("Contract", options=["FULL", "ON CALL", "CERT", "ASSIST"]),
                "Responsible": st.column_config.TextColumn("Responsible"),
                "Flight_Points": st.column_config.NumberColumn("Points", disabled=True),
                "LAE": st.column_config.TextColumn("LAE Code Assigned"),
                "Mech_1": st.column_config.TextColumn("Mech 1 Code Assigned"),
                "Mech_2": st.column_config.TextColumn("Mech 2 Code Assigned")
            }
        )
        
        for col in ["LAE", "Mech_1", "Mech_2", "Contract", "Responsible"]:
            if col in edited_flights.columns:
                st.session_state.flights_data[col] = edited_flights[col]
                
        updated_staff, updated_wl = recalculate_manual_workload(
            st.session_state.flights_data, 
            st.session_state.staff_data
        )
        st.session_state.staff_data = updated_staff
        st.session_state.workload_summary = updated_wl
    else:
        st.info("ℹ️ No schedule data available. Please upload the schedule file from the sidebar.")

# --- TAB 2: STAFF ROSTER & PRIVILEGES ---
with tab2:
    st.markdown("### 🛡️ Staff Roster & License Privileges")
    st.markdown("Review and manage technical authorizations and customer clearances.")
    
    if not st.session_state.staff_data.empty:
        cols_to_show_staff = ["Code", "Name", "Role", "Shift", "Time_IN", "Time_OUT", "Assignment", "Privileges", "Customer"]
        cols_exist_staff = [c for c in cols_to_show_staff if c in st.session_state.staff_data.columns]
        
        edited_staff = st.data_editor(
            st.session_state.staff_data[cols_exist_staff],
            num_rows="dynamic",
            use_container_width=True,
            key="staff_editor",
            column_config={
                "Code": st.column_config.TextColumn("Staff Code"),
                "Role": st.column_config.SelectboxColumn("Role", options=["LAE", "Mech"]),
                "Shift": st.column_config.TextColumn("SKD / Shift"),
                "Assignment": st.column_config.TextColumn("Assigned Flights"),
                "Privileges": st.column_config.TextColumn("Aircraft Authorizations"),
                "Customer": st.column_config.TextColumn("Customer Authorizations")
            }
        )
        for col in ["Code", "Role", "Shift", "Privileges", "Customer", "Assignment"]:
            if col in edited_staff.columns:
                st.session_state.staff_data[col] = edited_staff[col]
    else:
        st.info("ℹ️ No personnel data available. Please upload a schedule file.")

# --- TAB 3: WORKLOAD & SHIFT BALANCE ANALYSIS ---
with tab3:
    st.markdown("### 📊 Workload & Shift Balance Report")
    
    st.markdown("#### 👥 Workforce Roster Breakdown by Shift (SKD: D, N) - MECH & FLEX Groups")
    st.markdown("Listing personnel belonging to MECH and FLEX groups categorized by their scheduled shift (**D** or **N**) with automatic daily rotating **'Work Force'** assignment status.")
    
    if not st.session_state.staff_data.empty:
        staff_all = st.session_state.staff_data.copy()
        
        wf_records = []
        for _, s_row in staff_all.iterrows():
            s_code = str(s_row.get('Code', '')).strip().upper()
            s_name = str(s_row.get('Name', '')).strip().upper()
            s_role = str(s_row.get('Role', '')).strip().upper()
            s_shift = str(s_row.get('Shift', '')).strip().upper()
            s_assignment = str(s_row.get('Assignment', '')).strip()
            
            is_flex = s_code in LAE_FLEX_MECH_LIST or any(f_code in s_name for f_code in LAE_FLEX_MECH_LIST)
            is_mech = (s_role == 'MECH') or is_flex
            
            if is_mech:
                group_type = "FLEX Group" if is_flex else "Regular MECH"
                wf_records.append({
                    "Staff Code": s_code,
                    "Name": s_row.get('Name', ''),
                    "Group Category": group_type,
                    "Shift (SKD)": s_shift,
                    "Assignment Status": s_assignment if s_assignment else "-",
                    "Time IN": str(s_row.get('Time_IN', '-')),
                    "Time OUT": str(s_row.get('Time_OUT', '-'))
                })
        
        df_wf = pd.DataFrame(wf_records)
        if not df_wf.empty:
            df_dn = df_wf[df_wf['Shift (SKD)'].isin(['D', 'N'])].sort_values(by=['Shift (SKD)', 'Group Category', 'Staff Code'])
            
            c_wf1, c_wf2 = st.columns(2)
            with c_wf1:
                st.markdown("##### ☀️ Day Shift (SKD: D)")
                df_d_shift = df_dn[df_dn['Shift (SKD)'] == 'D']
                if not df_d_shift.empty:
                    st.dataframe(df_d_shift[['Staff Code', 'Name', 'Group Category', 'Assignment Status', 'Time IN', 'Time OUT']], use_container_width=True, hide_index=True)
                else:
                    st.info("No MECH or FLEX personnel found on 'D' shift.")
                    
            with c_wf2:
                st.markdown("##### 🌙 Night Shift (SKD: N)")
                df_n_shift = df_dn[df_dn['Shift (SKD)'] == 'N']
                if not df_n_shift.empty:
                    st.dataframe(df_n_shift[['Staff Code', 'Name', 'Group Category', 'Assignment Status', 'Time IN', 'Time OUT']], use_container_width=True, hide_index=True)
                else:
                    st.info("No MECH or FLEX personnel found on 'N' shift.")
        else:
            st.info("No MECH or FLEX personnel records identified.")
    else:
        st.info("Please load personnel data to view the workforce roster breakdown.")
        
    st.markdown("---")
    
    if st.session_state.workload_summary:
        c1, c2 = st.columns(2)
        
        with c1:
            st.markdown("#### ⚖️ Individual Workload Summary (Max Limit: 4.0 Pts)")
            wl_rows = []
            for code, points in st.session_state.workload_summary.items():
                s_row = st.session_state.staff_data[st.session_state.staff_data['Code'] == code]
                name = s_row['Name'].values[0] if not s_row.empty else code
                role = s_row['Role'].values[0] if not s_row.empty else '-'
                shift = s_row['Shift'].values[0] if not s_row.empty else '-'
                assigned = s_row['Assignment'].values[0] if not s_row.empty else ''
                status = "🟢 Normal" if points <= 4.0 else "🔴 Overloaded!"
                
                wl_rows.append({
                    "Code": code,
                    "Name": name,
                    "Role": role,
                    "Shift": shift,
                    "Assigned Flights / Status": assigned,
                    "Allocated Points": points,
                    "Status": status
                })
            df_wl_summary = pd.DataFrame(wl_rows)
            st.dataframe(df_wl_summary, use_container_width=True)
            
        with c2:
            st.markdown("#### 📈 Shift Balance Comparison (D vs N)")
            if not df_wl_summary.empty:
                shift_stats = df_wl_summary.groupby('Shift')['Allocated Points'].agg(['count', 'sum', 'mean']).reset_index()
                shift_stats.columns = ['Shift', 'Headcount', 'Total Points', 'Average Points / Person']
                st.dataframe(shift_stats, use_container_width=True)
                
            st.markdown("---")
            st.markdown("#### ⚠️ Unassigned Flights (Overtime Required)")
            shortages = []
            valid_lae_contracts = ['FULL', 'ON CALL', 'ONCALL', 'CERT']
            excluded_lae_flights = ['FM857', 'FM858', 'FM831', 'FM832']
            
            for idx, row in st.session_state.flights_data.iterrows():
                contract = str(row.get('Contract', '')).strip().upper()
                airline = str(row.get('Airline', '')).strip().upper()
                flt_no = str(row.get('Flight_No', '')).strip().upper()
                flt_id = f"{airline}{flt_no}".replace(" ", "")
                
                lae_blank = not str(row.get('LAE', '')).strip() or str(row.get('LAE', '')).strip().upper() == 'NONE'
                mech1_blank = not str(row.get('Mech_1', '')).strip() or str(row.get('Mech_1', '')).strip().upper() == 'NONE'
                mech2_blank = not str(row.get('Mech_2', '')).strip() or str(row.get('Mech_2', '')).strip().upper() == 'NONE'
                
                req_mech = 2 if airline == 'TG' else (1 if airline == 'TR' else (2 if contract in ['FULL', 'ASSIST', 'CERT'] else 1))
                is_excluded_flight = any(ex in flt_id for ex in excluded_lae_flights)
                missing_roles = []
                if ((contract in valid_lae_contracts) or (airline == 'TR')) and (not is_excluded_flight) and lae_blank: 
                    missing_roles.append("LAE")
                if mech1_blank: 
                    missing_roles.append("Mech 1")
                if req_mech == 2 and mech2_blank: 
                    missing_roles.append("Mech 2")
                if missing_roles:
                    shortages.append({
                        "Flight": row.get('Flight_ID', '-'),
                        "A/C Type": row.get('A/C Type', '-'),
                        "Contract": row.get('Contract', '-'),
                        "STA / STD": f"{row.get('STA', '')} - {row.get('STD', '')}",
                        "Missing Roles (OT Needed)": ", ".join(missing_roles)
                    })
            
            if shortages:
                st.warning(f"⚠️ Total incomplete flights found: {len(shortages)}")
                st.dataframe(pd.DataFrame(shortages), use_container_width=True)
            else:
                st.success("🎉 All flights have been fully and successfully allocated!")
    else:
        st.info("💡 Click 'Run Auto-Allocation' in the first tab to calculate and view workload reports.")